import io
import os
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, or_

from app.database import get_db
from app.models.task import Task, TaskStatus, SearchMode
from app.models.lead import Lead, LeadCategory
from app.models.platform_result import PlatformResult
from app.platforms.registry import PlatformRegistry
from app.services.task_engine import TaskEngine, get_task_engine
from app.services.keyword_expander import KeywordExpander
from app.services.llm_client import LLMClient
from app.services.report_generator import ReportGenerator
from app.services.cost_engine import CostEngine
from app.schemas import (
    CreateTaskRequest, TaskBrief, TaskDetail, LeadListResponse, LeadBrief,
    PlatformInfo, PlatformResultBrief, SearchModeOption,
)
from app.platforms import reddit_platform, google_search, youtube_platform, tier1_stubs  # noqa: F401 - trigger registry

router = APIRouter()


SEARCH_MODES: List[SearchModeOption] = [
    SearchModeOption(value="customers", label="找客户", description="寻找有购买意向的潜在客户"),
    SearchModeOption(value="companies", label="找企业", description="发现有相关需求的公司"),
    SearchModeOption(value="demands", label="找需求", description="挖掘市场正在寻找的产品/服务"),
    SearchModeOption(value="suppliers", label="找供应商", description="匹配潜在的供应商资源"),
    SearchModeOption(value="partners", label="找合作", description="识别潜在的合作对象"),
    SearchModeOption(value="opportunities", label="找市场机会", description="发现新兴的商业机会和蓝海"),
    SearchModeOption(value="family_bucket", label="全家桶（推荐）", description="综合模式：同时搜索客户、企业、需求、机会等所有维度"),
]


@router.get("/meta/search-modes", response_model=List[SearchModeOption])
def list_search_modes():
    return SEARCH_MODES


@router.get("/meta/platforms", response_model=List[PlatformInfo])
def list_platforms():
    return PlatformRegistry.list_all()


@router.get("/meta/lead-categories")
def list_lead_categories():
    return [{"value": c.value, "label": {
        "customer": "客户", "company": "企业", "demand": "需求", "pain_point": "痛点",
        "opportunity": "机会", "trend": "趋势", "competition": "竞争",
    }[c.value]} for c in LeadCategory]


@router.post("/keywords/expand")
async def expand_keywords(query: str = Query(..., min_length=1, max_length=500),
                          search_mode: SearchMode = SearchMode.FAMILY_BUCKET):
    expander = KeywordExpander(LLMClient())
    keywords = await expander.expand(query, search_mode.value)
    return {"query": query, "search_mode": search_mode.value, "keywords": keywords}


@router.post("/tasks", response_model=TaskBrief)
def create_task(req: CreateTaskRequest, db: Session = Depends(get_db)):
    task = TaskEngine.create_task(db, req.query, req.search_mode, req.platforms)
    engine = get_task_engine()
    engine.start_task(task.id)
    return task


@router.get("/tasks", response_model=List[TaskBrief])
def list_tasks(db: Session = Depends(get_db), limit: int = 50, offset: int = 0):
    tasks = db.query(Task).order_by(desc(Task.created_at)).limit(limit).offset(offset).all()
    return tasks


@router.get("/tasks/{task_id}", response_model=TaskDetail)
def get_task(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    platforms = db.query(PlatformResult).filter(PlatformResult.task_id == task_id).order_by(PlatformResult.id).all()
    cost_summary = CostEngine.summarize_task(db, task_id)
    return TaskDetail(
        task=TaskBrief.model_validate(task, from_attributes=True),
        platforms=[PlatformResultBrief.model_validate(p, from_attributes=True) for p in platforms],
        expanded_keywords=task.expanded_keywords or [],
        cost_summary=cost_summary,
    )


@router.post("/tasks/{task_id}/start", response_model=TaskBrief)
def start_task(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    task.status = TaskStatus.PENDING
    task.error_message = None
    db.commit()
    engine = get_task_engine()
    engine.start_task(task_id)
    return task


@router.get("/tasks/{task_id}/leads", response_model=LeadListResponse)
def list_task_leads(
    task_id: int,
    category: Optional[LeadCategory] = None,
    high_intent_only: bool = False,
    min_score: float = 0.0,
    platform: Optional[str] = None,
    keyword: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    q = db.query(Lead).filter(Lead.task_id == task_id)
    if category:
        q = q.filter(Lead.category == category)
    if high_intent_only:
        q = q.filter(Lead.is_high_intent == 1)
    if min_score > 0:
        q = q.filter(Lead.overall_score >= min_score)
    if platform:
        q = q.filter(Lead.source_platform == platform)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(or_(Lead.title.like(kw), Lead.summary.like(kw), Lead.content.like(kw), Lead.company_name.like(kw)))
    total = q.count()
    high_total = q.filter(Lead.is_high_intent == 1).count() if not high_intent_only else total
    items = q.order_by(desc(Lead.overall_score)).limit(limit).offset(offset).all()
    return LeadListResponse(total=total, high_intent_total=high_total, items=[LeadBrief.model_validate(x, from_attributes=True) for x in items])


@router.get("/tasks/{task_id}/report.md")
async def get_task_report_md(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    gen = ReportGenerator(db, task)
    md = await gen.generate_markdown()
    return StreamingResponse(io.BytesIO(md.encode("utf-8")), media_type="text/markdown; charset=utf-8",
                             headers={"Content-Disposition": f"attachment; filename=\"report_task_{task_id}.md\""})


@router.get("/tasks/{task_id}/report.html")
async def get_task_report_html(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    gen = ReportGenerator(db, task)
    md = await gen.generate_markdown()
    try:
        import markdown
        body = markdown.markdown(md, extensions=["tables", "fenced_code"])
    except Exception:
        body = "<pre>" + md.replace("<", "&lt;") + "</pre>"
    html = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>报告 - {task.query}</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;max-width:960px;margin:30px auto;padding:0 20px;line-height:1.7;color:#222}}
h1,h2,h3{{color:#111;margin-top:2em}}
h1{{border-bottom:3px solid #2563eb;padding-bottom:.3em}}
h2{{border-bottom:1px solid #d1d5db;padding-bottom:.2em}}
table{{border-collapse:collapse;width:100%;margin:1em 0;font-size:14px}}
th,td{{border:1px solid #e5e7eb;padding:8px 10px;text-align:left;vertical-align:top}}
th{{background:#f3f4f6;font-weight:600}}
blockquote{{border-left:4px solid #2563eb;margin:1em 0;padding:.5em 1em;background:#f8fafc;color:#475569}}
code{{background:#f1f5f9;padding:2px 6px;border-radius:4px}}
.hi{{background:#fef3c7}}
</style></head><body>{body}</body></html>"""
    return HTMLResponse(content=html)


@router.get("/tasks/{task_id}/export.xlsx")
def export_leads_xlsx(
    task_id: int,
    high_intent_only: bool = True,
    min_score: float = 0.0,
    db: Session = Depends(get_db),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "Task not found")
    q = db.query(Lead).filter(Lead.task_id == task_id)
    if high_intent_only:
        q = q.filter(Lead.is_high_intent == 1)
    if min_score > 0:
        q = q.filter(Lead.overall_score >= min_score)
    leads = q.order_by(desc(Lead.overall_score)).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Leads_{task_id}"
    headers = ["#", "类别", "综合分", "意向分", "机会分", "紧急分", "高意向",
               "标题/摘要", "公司", "作者", "地点", "语言", "平台", "URL", "原文链接",
               "标签", "发布时间", "分析说明", "推荐行动"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    category_label = {"customer": "客户", "company": "企业", "demand": "需求", "pain_point": "痛点",
                      "opportunity": "机会", "trend": "趋势", "competition": "竞争"}
    hi_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    for i, l in enumerate(leads, 1):
        row = [
            i,
            category_label.get(l.category.value, l.category.value),
            l.overall_score, l.intent_score, l.opportunity_score, l.urgency_score,
            "是" if l.is_high_intent else "否",
            (l.title or l.summary or "")[:200],
            l.company_name or "", l.author or "", l.location or "", l.language or "",
            l.source_platform or "",
            l.url or "",
            ", ".join((l.evidence_links or [])[:5]),
            ", ".join((l.tags or [])[:20]),
            l.published_at.strftime("%Y-%m-%d %H:%M") if l.published_at else "",
            (l.analysis_notes or "")[:500],
            "；".join((l.recommendations or [])[:5]),
        ]
        ws.append(row)
        if l.is_high_intent:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = hi_fill
    for col in [8, 14, 18, 19]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 50
    for col in [9, 10, 13, 16]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Task_Info")
    info.append(["字段", "值"])
    info.append(["任务ID", task.id])
    info.append(["搜索主题", task.query])
    info.append(["搜索模式", task.search_mode.value])
    info.append(["平台", ", ".join(task.platforms or [])])
    info.append(["扩展关键词数", len(task.expanded_keywords or [])])
    info.append(["状态", task.status.value])
    info.append(["创建时间", task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else ""])
    info.append(["完成时间", task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if task.completed_at else ""])
    info.append(["总内容", task.total_contents])
    info.append(["高意向线索", task.total_high_intent_leads])
    info.append(["企业线索", task.total_companies])
    info.append(["需求", task.total_demands])
    info.append(["机会", task.total_company_opportunities])
    info.append(["趋势", task.total_trends])
    info.append(["竞争", task.total_competitions])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"leads_task_{task_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{fname}\""}
    )
